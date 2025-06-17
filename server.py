from flask import Flask, request, jsonify
from flask_cors import CORS
import requests
import os
import time
import logging
import base64
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import warnings
import json
from datetime import datetime
from typing import List, Dict, Any
import re

# Suppress the WMF image warning
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# Configure logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app)

# WordPress API credentials, change this to your own credentials
WP_USER = "alif.adrian"
WP_APP_PASSWORD = "xBa8 zJjb UDhK oXcP NeqR 7MDm"
WP_API_URL = "https://frontier.vstecs.ai/wp-json/wp/v2/pages"

logger.info("WordPress credentials loaded successfully")

def is_cell_bold(cell):
    """Helper function to safely check if a cell's font is bold"""
    try:
        if not cell or not hasattr(cell, 'font'):
            return False
        return cell.font.bold if cell.font else False
    except Exception as e:
        logger.debug(f"Error checking bold: {str(e)}")
        return False

def get_cell_value_safe(cell):
    """Helper function to safely get cell value"""
    try:
        if not cell:
            return ''
        return str(cell.value) if cell.value is not None else ''
    except Exception as e:
        logger.debug(f"Error getting cell value: {str(e)}")
        return ''

def get_cell_number_format(cell):
    """Helper function to safely get cell number format"""
    try:
        if not cell:
            return ''
        return str(cell.number_format) if cell.number_format else ''
    except Exception as e:
        logger.debug(f"Error getting number format: {str(e)}")
        return ''

def convert_to_rag_format(table_data: List[Dict[str, Any]], use_ollama: bool = True) -> str:
    """
    Convert table data to RAG format with enhanced searchability and consistent structure
    """
    try:
        if use_ollama:
            try:
                # Prepare the prompt
                prompt = f"""
                Convert each row of the input table into exactly this format:
                PRODUCT IDENTIFIERS:
                - Product Number: {{product_number}}
                - Series: {{series}}
                - Machine Type Model (MTM): {{mtm}}
                - SKU: {{sku}}
                - Model: {{model}}
                - Part Number: {{part_number}}

                PRODUCT DETAILS:
                Description: {{description}}
                Warranty: {{warranty}}
                Order Reason Code: {{order_reason}}
                Special Conditions: {{special_conditions}}
                Additional Information: {{additional_info}}

                PRICING INFORMATION:
                Global List Price: {{global_price}}
                Disti Price: {{disti_price}}
                Margin%: {{margin}}
                Reseller Price: {{reseller_price}}
                Suggested Retail Price (SRP): {{srp}}
                SRP Inc Tax: {{srp_tax}}

                Rules:
                1. For Product Identifiers:
                   - Extract ALL possible identifiers from the data
                   - If a field is empty, use "N/A"
                   - For MTM, extract from product number if not explicitly provided
                   - Include both original and formatted versions of numbers
                   
                2. For Product Details:
                   - Include only non-empty fields
                   - Combine all relevant product details
                   - Remove any empty or redundant information
                   - Format in a clear, structured way
                   
                3. For Pricing Information:
                   - Include all available price fields
                   - Format prices with currency (USD/SGD)
                   - Include margin percentage if available
                   - Include tax-inclusive price if available
                   - Use "N/A" for missing values
                   
                4. Additional Rules:
                   - Skip any empty rows
                   - Remove any columns that are completely empty
                   - Format each product entry with clear line breaks
                   - Use consistent formatting across all entries
                   - Include ALL possible search terms that could identify the product
                   - Preserve all warranty information
                   - Include order reason codes and special conditions
                   - Add any additional information provided

                Input table data:
                {json.dumps(table_data, indent=2)}

                Example output for Format 1 (Simple):
                PRODUCT IDENTIFIERS:
                - Product Number: SS-CV-ENT-1M
                - Series: SS-CV
                - Machine Type Model (MTM): SS-CV
                - SKU: SS-CV-ENT-1M
                - Model: SS-CV-ENT-1M
                - Part Number: 1054995

                PRODUCT DETAILS:
                Description: Enterprise License for 1 Month
                Warranty: N/A
                Order Reason Code: N/A
                Special Conditions: N/A
                Additional Information: N/A

                PRICING INFORMATION:
                Global List Price: USD 500
                Disti Price: N/A
                Margin%: N/A
                Reseller Price: N/A
                Suggested Retail Price (SRP): N/A
                SRP Inc Tax: N/A

                Example output for Format 2 (Detailed):
                PRODUCT IDENTIFIERS:
                - Product Number: 12U8005PSG
                - Series: ThinkCentre
                - Machine Type Model (MTM): 12U8
                - SKU: 12U8005PSG
                - Model: ThinkCentre M70s Gen 5
                - Part Number: 12U8-005

                PRODUCT DETAILS:
                Description: ThinkCentre M70s Gen 5: 260W SFF (Q670 Chipset) - Intel® Core™ i5-14400, 10C (6P + 4E) / 16T, P-core 2.5 / 4.7GHz, E-core 1.8 / 3.5GHz, 20MB - 8GB UDIMM DDR5-4800 - 512GB SSD M.2 2280 PCIe® 4.0x4 Performance NVMe® Opal 2.0 / No ODD / 2.5" HDD Bracket Kit / Internal Speaker - Graphics: Integrated Intel® UHD Graphics 730 - Intel® Wi-Fi® 6E AX211, 802.11ax 2x2 + BT5.3, vPro® - USB Keyboard / Mouse - HDMI / Display Port / LAN - Windows® 11 Pro, English / NO Recovery Media - 3 Year on-site
                Warranty: 3 Year on-site
                Order Reason Code: N/A
                Special Conditions: N/A
                Additional Information: N/A

                PRICING INFORMATION:
                Global List Price: N/A
                Disti Price: USD 1,200
                Margin%: 15%
                Reseller Price: USD 1,380
                Suggested Retail Price (SRP): USD 1,399
                SRP Inc Tax: USD 1,499
                """

                # Call OLLAMA API with timeout
                ollama_url = "http://172.18.0.221:7870/api/generate"
                response = requests.post(
                    ollama_url,
                    json={
                        "model": "gemma:3.27b",
                        "prompt": prompt,
                        "stream": False
                    },
                    timeout=30  # 30 seconds timeout
                )

                if response.status_code == 200:
                    return response.json().get("response", "")
                else:
                    logger.error(f"OLLAMA API error: {response.text}")
                    raise Exception(f"OLLAMA API error: {response.text}")
            except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as e:
                logger.warning(f"OLLAMA connection error, falling back to simple conversion: {str(e)}")
                use_ollama = False
            except Exception as e:
                logger.error(f"Error calling OLLAMA: {str(e)}")
                use_ollama = False

        # Fallback to simple conversion if OLLAMA is not available
        if not use_ollama:
            results = []
            for row in table_data:
                # Skip empty rows
                if not any(str(value).strip() for value in row.values()):
                    continue

                # Extract all possible identifiers
                identifiers = {
                    'product_number': "N/A",
                    'series': "N/A",
                    'mtm': "N/A",
                    'sku': "N/A",
                    'model': "N/A",
                    'part_number': "N/A"
                }

                # Find all possible identifiers
                for key, value in row.items():
                    if isinstance(value, str) and value.strip():
                        key_lower = str(key).lower().replace('#', '')
                        value = str(value).strip()
                        
                        if any(term in key_lower for term in ['product', 'code', 'number']):
                            identifiers['product_number'] = value
                        if 'series' in key_lower:
                            identifiers['series'] = value
                        if 'sku' in key_lower:
                            identifiers['sku'] = value
                        if 'model' in key_lower:
                            identifiers['model'] = value
                        if 'part' in key_lower:
                            identifiers['part_number'] = value
                        if 'mtm' in key_lower:
                            identifiers['mtm'] = value

                # Extract MTM from product number if not found
                if identifiers['mtm'] == "N/A" and identifiers['product_number'] != "N/A":
                    mtm_match = re.match(r'^([A-Z0-9]{4,7})', identifiers['product_number'])
                    if mtm_match:
                        identifiers['mtm'] = mtm_match.group(1)

                # Combine non-empty fields for description
                description_parts = []
                for key, value in row.items():
                    if isinstance(value, (str, int, float)) and str(value).strip():
                        key_lower = str(key).lower().replace('#', '')
                        if key_lower not in ['code', 'model', 'part', 'sku', 'mtm', 'series', 'price', 'srp', 'disti', 'reseller', 'usd', 'warranty', 'margin', 'order', 'special', 'additional', 'global']:
                            description_parts.append(f"{key}: {str(value).strip()}")

                description = " - ".join(description_parts) if description_parts else "N/A"

                # Extract additional fields
                warranty = row.get('Warranty', 'N/A')
                order_reason = row.get('Order Reason Code', 'N/A')
                special_conditions = row.get('Special Conditions', 'N/A')
                additional_info = row.get('Additional Information', 'N/A')

                # Format prices
                def format_price(value, prefix=''):
                    if isinstance(value, (int, float)):
                        return f"{prefix} {value:,.2f}"
                    elif isinstance(value, str) and value.strip():
                        return value.strip()
                    return 'N/A'

                global_price = format_price(row.get('USD Global List Price', 'N/A'), 'USD')
                disti_price = format_price(row.get('Disti Price', 'N/A'), 'USD')
                margin = format_price(row.get('Margin%', 'N/A'), '')
                reseller_price = format_price(row.get('Reseller Price', 'N/A'), 'USD')
                srp = format_price(row.get('Suggested Retail Price (SRP)', 'N/A'), 'USD')
                srp_tax = format_price(row.get('SRP Inc Tax', 'N/A'), 'USD')

                # Format the output
                result = f"""PRODUCT IDENTIFIERS:
- Product Number: {identifiers['product_number']}
- Series: {identifiers['series']}
- Machine Type Model (MTM): {identifiers['mtm']}
- SKU: {identifiers['sku']}
- Model: {identifiers['model']}
- Part Number: {identifiers['part_number']}

PRODUCT DETAILS:
Description: {description}
Warranty: {warranty}
Order Reason Code: {order_reason}
Special Conditions: {special_conditions}
Additional Information: {additional_info}

PRICING INFORMATION:
Global List Price: {global_price}
Disti Price: {disti_price}
Margin%: {margin}
Reseller Price: {reseller_price}
Suggested Retail Price (SRP): {srp}
SRP Inc Tax: {srp_tax}

"""
                results.append(result)

            return "\n".join(results)

    except Exception as e:
        logger.error(f"Error in convert_to_rag_format: {str(e)}")
        return f"Error: {str(e)}"

def analyze_column_importance(headers: List[str], sample_data: List[Dict[str, Any]]) -> Dict[str, float]:
    """
    Analyze the importance of each column based on its content and characteristics
    """
    importance_scores = {}
    
    for header in headers:
        score = 0.0
        header_lower = str(header).lower()
        
        # Check column name patterns
        if any(term in header_lower for term in ['id', 'code', 'number', 'reference', 'sku']):
            score += 3.0  # High importance for identifiers
        elif any(term in header_lower for term in ['name', 'title', 'product', 'item']):
            score += 3.0  # High importance for names
        elif any(term in header_lower for term in ['description', 'detail', 'info', 'note']):
            score += 2.5  # High importance for descriptions
        elif any(term in header_lower for term in ['price', 'cost', 'amount', '$']):
            score += 2.5  # High importance for prices
        elif any(term in header_lower for term in ['date', 'time', 'period']):
            score += 1.5  # Medium importance for dates
        elif any(term in header_lower for term in ['status', 'state', 'condition']):
            score += 1.5  # Medium importance for status
        elif any(term in header_lower for term in ['category', 'type', 'group', 'class']):
            score += 2.0  # High importance for categories
        elif any(term in header_lower for term in ['quantity', 'amount', 'count']):
            score += 1.5  # Medium importance for quantities
        
        # Analyze content patterns
        non_empty_values = [row.get(header, '') for row in sample_data if row.get(header, '')]
        if non_empty_values:
            # Check for unique values
            unique_ratio = len(set(non_empty_values)) / len(non_empty_values)
            score += unique_ratio * 1.0  # Higher score for more unique values
            
            # Check for numeric values
            numeric_count = sum(1 for v in non_empty_values if str(v).replace('.', '').replace(',', '').isdigit())
            if numeric_count / len(non_empty_values) > 0.5:
                score += 1.0  # Bonus for numeric columns
            
            # Check for date-like values
            date_pattern = r'\d{1,4}[-/]\d{1,2}[-/]\d{1,4}'
            date_count = sum(1 for v in non_empty_values if re.search(date_pattern, str(v)))
            if date_count / len(non_empty_values) > 0.5:
                score += 0.5  # Bonus for date columns
        
        importance_scores[header] = score
    
    return importance_scores

def generate_contexts_from_excel(table_data: List[Dict[str, Any]], sheet_name: str) -> List[Dict[str, Any]]:
    """
    Automatically generate contexts from Excel data by analyzing the structure and content
    """
    if not table_data:
        return []
        
    # Get all unique headers from the data
    headers = set()
    for row in table_data:
        headers.update(row.keys())
    headers = list(headers)
    
    # Analyze column importance
    importance_scores = analyze_column_importance(headers, table_data)
    
    # Sort headers by importance
    sorted_headers = sorted(headers, key=lambda h: importance_scores.get(h, 0), reverse=True)
    
    contexts = []
    for row in table_data:
        if not any(str(value).strip() for value in row.values()):
            continue
            
        # Initialize context with basic structure
        context = {
            'category': sheet_name,
            'metadata': {
                'column_importance': {h: importance_scores.get(h, 0) for h in headers},
                'identifiers': {},
                'descriptions': {},
                'measurements': {},
                'dates': {},
                'status': {},
                'relationships': {},
                'additional': {}
            }
        }
        
        # Process each column based on its importance and content
        for header in sorted_headers:
            value = row.get(header, '')
            if not value or str(value).strip() == '':
                continue
                
            value_str = str(value).strip()
            header_lower = str(header).lower()
            
            # Categorize the column based on its name and content
            if any(term in header_lower for term in ['id', 'code', 'number', 'reference', 'sku']):
                context['metadata']['identifiers'][header] = value_str
                if not context.get('product_number'):
                    context['product_number'] = value_str
            elif any(term in header_lower for term in ['name', 'title', 'product', 'item']):
                context['product_name'] = value_str
            elif any(term in header_lower for term in ['description', 'detail', 'info', 'note']):
                context['metadata']['descriptions'][header] = value_str
                if not context.get('description'):
                    context['description'] = value_str
            elif any(term in header_lower for term in ['price', 'cost', 'amount', '$']):
                try:
                    context['price'] = float(str(value).replace('$', '').replace(',', ''))
                except:
                    context['price'] = value_str
            elif any(term in header_lower for term in ['date', 'time', 'period']):
                context['metadata']['dates'][header] = value_str
            elif any(term in header_lower for term in ['status', 'state', 'condition']):
                context['metadata']['status'][header] = value_str
            elif any(term in header_lower for term in ['quantity', 'amount', 'count', 'size', 'weight', 'dimension']):
                context['metadata']['measurements'][header] = value_str
            elif any(term in header_lower for term in ['category', 'type', 'group', 'class']):
                context['category'] = value_str
            elif any(term in header_lower for term in ['related', 'parent', 'child', 'component']):
                context['metadata']['relationships'][header] = value_str
            else:
                # Store any other column as additional metadata
                context['metadata']['additional'][header] = value_str
        
        # Ensure required fields exist
        if 'product_name' not in context:
            # Try to find a suitable name from identifiers or descriptions
            for identifier in context['metadata']['identifiers'].values():
                if len(identifier) > 3:  # Avoid using very short identifiers as names
                    context['product_name'] = identifier
                    break
            if 'product_name' not in context:
                context['product_name'] = f"Item {len(contexts) + 1}"
        
        if 'description' not in context:
            # Combine all descriptions if available
            descriptions = list(context['metadata']['descriptions'].values())
            if descriptions:
                context['description'] = ' | '.join(descriptions)
            else:
                context['description'] = context['product_name']
        
        # Clean up empty metadata
        for key in list(context['metadata'].keys()):
            if not context['metadata'][key]:
                del context['metadata'][key]
        
        # Add column analysis information
        context['metadata']['column_analysis'] = {
            'total_columns': len(headers),
            'important_columns': [h for h in sorted_headers if importance_scores.get(h, 0) > 2.0],
            'column_types': {
                'identifiers': list(context['metadata'].get('identifiers', {}).keys()),
                'descriptions': list(context['metadata'].get('descriptions', {}).keys()),
                'measurements': list(context['metadata'].get('measurements', {}).keys()),
                'dates': list(context['metadata'].get('dates', {}).keys()),
                'status': list(context['metadata'].get('status', {}).keys())
            }
        }
        
        contexts.append(context)
    
    return contexts

def extract_specifications(description: str) -> Dict[str, Any]:
    """
    Extract specifications from product description using enhanced pattern matching
    """
    specs = {}
    
    # Enhanced specification patterns
    patterns = {
        'processor': r'(?:Intel|AMD|Core|Ryzen|i\d|i\d-\d{4}[A-Z]?|Snapdragon|MediaTek|Apple\s+[A-Z0-9]+)',
        'ram': r'(\d+GB(?:\s+RAM|\s+DDR\d)?)',
        'storage': r'(\d+GB(?:\s+SSD|\s+HDD|\s+NVMe)?)',
        'display': r'(\d+(?:\.\d+)?["\'](?:\s+FHD|\s+UHD|\s+4K|\s+Retina|\s+OLED|\s+LCD)?)',
        'os': r'(Windows\s+\d+(?:\s+Pro|\s+Home|\s+Enterprise)?|Linux|macOS|iOS|Android)',
        'warranty': r'(\d+\s+Year(?:\s+on-site|\s+carry-in|\s+limited)?)',
        'battery': r'(\d+(?:\.\d+)?\s*(?:mAh|Wh|hours?))',
        'resolution': r'(\d+x\d+(?:\s+p)?)',
        'refresh_rate': r'(\d+(?:\.\d+)?\s*Hz)',
        'ports': r'(USB\s+\d+(?:\.\d+)?|Thunderbolt\s+\d+|HDMI\s+\d+(?:\.\d+)?|DisplayPort\s+\d+(?:\.\d+)?)',
        'network': r'(Wi-Fi\s+\d+(?:\.\d+)?|Bluetooth\s+\d+(?:\.\d+)?|5G|4G|LTE)',
        'camera': r'(\d+(?:\.\d+)?\s*MP(?:\s+camera)?)',
        'weight': r'(\d+(?:\.\d+)?\s*(?:kg|g|lbs))',
        'dimensions': r'(\d+(?:\.\d+)?\s*(?:mm|cm|inch)(?:\s*x\s*\d+(?:\.\d+)?\s*(?:mm|cm|inch)){2})'
    }
    
    # Extract specifications using patterns
    for key, pattern in patterns.items():
        matches = re.finditer(pattern, description, re.IGNORECASE)
        values = [match.group(0) for match in matches]
        if values:
            specs[key] = values[0] if len(values) == 1 else values
    
    # Extract version numbers
    version_pattern = r'(?:v|version|ver\.?)\s*(\d+(?:\.\d+){1,3})'
    version_matches = re.finditer(version_pattern, description, re.IGNORECASE)
    versions = [match.group(1) for match in version_matches]
    if versions:
        specs['version'] = versions[0] if len(versions) == 1 else versions
    
    # Extract model numbers
    model_pattern = r'(?:model|type)\s*(?:number|no\.?)?\s*[:#]?\s*([A-Z0-9-]+)'
    model_matches = re.finditer(model_pattern, description, re.IGNORECASE)
    models = [match.group(1) for match in model_matches]
    if models:
        specs['model_number'] = models[0] if len(models) == 1 else models
    
    return specs

def convert_excel_to_html(file_path):
    try:
        logger.debug(f"Loading workbook: {file_path}")
        wb = load_workbook(file_path, data_only=True)
        
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Initialize context storage
        product_contexts = []
        product_groups = {}
        current_category = None
        current_product_name = None
        
        # CSS styles as a separate string
        css_styles = """
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700&display=swap');
            
            body { 
                font-family: 'Poppins', sans-serif; 
                line-height: 1.6; 
                color: #333; 
                margin: 0; 
                padding: 20px; 
                background-color: #f4f4f4; 
                min-height: 100vh;
                display: flex;
                flex-direction: column;
            }
            h1, h2, h3 { color: #2c3e50; margin-bottom: 20px; }
            .category-header { background-color: #f8f9fa; padding: 15px; margin: 20px 0; border-left: 5px solid #2c3e50; font-size: 1.2em; font-weight: bold; }
            .main-title { font-size: 2em; text-align: center; margin-bottom: 30px; color: #2c3e50; }
            .subtitle { font-size: 1.5em; text-align: center; margin: 20px 0; color: #34495e; }
            table { width: 100%; border-collapse: collapse; margin: 20px 0; background-color: #fff; border: 1px solid #ddd; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
            th, td { border: 1px solid #ddd; padding: 12px 15px; text-align: left; vertical-align: middle; }
            th { background-color: #f8f9fa; font-weight: bold; color: #2c3e50; border-bottom: 2px solid #ddd; }
            td { border-bottom: 1px solid #ddd; }
            tr:nth-child(even) { background-color: #f9f9f9; }
            tr:hover { background-color: #f5f5f5; }
            .empty-row { height: 20px; background-color: transparent; border: none; }
            .empty-row td { border: none; }
            .price-column { text-align: right; font-family: monospace; white-space: nowrap; }
            .product-number-column { font-family: monospace; white-space: nowrap; }
            .description-column { min-width: 300px; }
            .section-header { background-color: #f8f9fa; font-weight: bold; color: #2c3e50; }
            .section-header td { padding: 15px; font-size: 1.1em; }
            .metadata { text-align: center; color: #666; margin: 10px 0; font-style: italic; }
            .context-preview { 
                background-color: #fff; 
                border: 1px solid #ddd; 
                border-radius: 4px; 
                padding: 20px; 
                margin: 20px 0; 
                box-shadow: 0 2px 4px rgba(0,0,0,0.1); 
                display: none; /* Hidden by default */
            }
            .context-item {
                border: 1px solid #eee;
                padding: 15px;
                margin: 10px 0;
                border-radius: 4px;
                background-color: #f8f9fa;
            }
            .context-item h4 {
                margin: 0 0 10px 0;
                color: #2c3e50;
            }
            .context-item pre {
                background-color: #fff;
                padding: 10px;
                border-radius: 4px;
                overflow-x: auto;
                margin: 0;
            }
            .context-summary {
                background-color: #e8f4f8;
                padding: 15px;
                margin: 10px 0;
                border-radius: 4px;
                border-left: 4px solid #2c3e50;
            }
            .context-summary h4 {
                margin: 0 0 10px 0;
                color: #2c3e50;
            }
            .context-summary ul {
                margin: 0;
                padding-left: 20px;
            }
            .context-summary li {
                margin: 5px 0;
            }
            .show-context {
                display: block !important;
            }
            .convert-button {
                background-color: #2c3e50;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 4px;
                cursor: pointer;
                font-size: 1em;
                margin: 20px auto;
                display: block;
            }
            .convert-button:hover {
                background-color: #34495e;
            }
        </style>
        """
        
        html_content = f"""
        {css_styles}
        <h1 class="main-title">Price List</h1>
        <div class="metadata">
            <p>Last Updated: {current_time}</p>
        </div>
        """
        
        # Store table data for RAG conversion
        table_data = []
        
        for sheet in wb.sheetnames:
            logger.debug(f"Processing sheet: {sheet}")
            ws = wb[sheet]
            
            html_content += f'<h2 class="subtitle">{sheet}</h2>\n'
            html_content += '<div class="table-container">\n<table>\n'
            
            is_header_section = False
            headers = {}
            first_row = next(ws.rows)
            
            # Process headers
            for idx, cell in enumerate(first_row):
                header_value = get_cell_value_safe(cell)
                if header_value:
                    headers[idx] = str(header_value).strip()
            
            # Add header row to HTML
            html_content += "<tr>"
            for idx in range(len(headers)):
                header = headers.get(idx, '')
                html_content += f'<th>{header}</th>'
            html_content += "</tr>\n"
            
            sheet_data = []
            for row_idx, row in enumerate(ws.rows):
                if row_idx == 0:  # Skip header row
                    continue
                    
                if all(get_cell_value_safe(cell) == '' for cell in row):
                    if not is_header_section:
                        html_content += '<tr class="empty-row"><td colspan="100%">&nbsp;</td></tr>\n'
                    continue
                
                first_cell = row[0]
                if is_cell_bold(first_cell) and get_cell_value_safe(first_cell):
                    current_category = get_cell_value_safe(first_cell)
                    current_product_name = current_category
                    is_header_section = True
                    html_content += f'<tr class="section-header"><td colspan="100%">{current_category}</td></tr>\n'
                    continue
                
                is_header_section = False
                html_content += "<tr>"
                
                row_data = {}
                for col_idx, cell in enumerate(row):
                    value = get_cell_value_safe(cell)
                    number_format = get_cell_number_format(cell)
                    
                    header = headers.get(col_idx, '')
                    if col_idx == 0:
                        css_class = 'product-number-column'
                    elif any(price_term in str(header).lower() for price_term in ['price', 'usd', '$']):
                        css_class = 'price-column'
                    elif col_idx == 1:
                        css_class = 'description-column'
                    else:
                        css_class = ''
                    
                    if value == '':
                        display_value = ''
                    elif isinstance(value, (int, float)):
                        if any(price_term in str(header).lower() for price_term in ['price', 'usd', '$']) or '$' in number_format:
                            display_value = f"${value:,.0f}"
                        elif ',' in number_format or '.' in number_format:
                            display_value = f"{value:,.0f}"
                        else:
                            display_value = str(value)
                    else:
                        display_value = str(value)
                    
                    if header:
                        row_data[header] = value
                    
                    html_content += f'<td class="{css_class}">{display_value}</td>'
                
                html_content += "</tr>\n"
                
                if row_data:
                    sheet_data.append(row_data)
                    table_data.append(row_data)
            
            html_content += "</table>\n</div>\n"
            
            # Generate contexts for this sheet
            if sheet_data:
                sheet_contexts = generate_contexts_from_excel(sheet_data, sheet)
                product_contexts.extend(sheet_contexts)
                
                # Add context preview for this sheet (hidden by default)
                html_content += f"""
                <div class="context-preview" id="context-preview-{sheet}">
                    <h3>Generated Contexts for {sheet}</h3>
                    <div class="context-summary">
                        <h4>Sheet Summary:</h4>
                        <ul>
                            <li>Total Products: {len(sheet_contexts)}</li>
                            <li>Categories: {', '.join(set(ctx.get('category', '') for ctx in sheet_contexts))}</li>
                            <li>Columns Used: {', '.join(headers.values())}</li>
                        </ul>
                    </div>
                    <div class="context-items">
                """
                
                for context in sheet_contexts:
                    # Ensure product_name exists
                    if 'product_name' not in context:
                        context['product_name'] = context.get('product_number', 'Unknown Product')
                    
                    html_content += f"""
                    <div class="context-item">
                        <h4>{context['product_name']}</h4>
                        <pre>{json.dumps(context, indent=2)}</pre>
                    </div>
                    """
                
                html_content += """
                    </div>
                </div>
                """
        
        # Add convert button and JavaScript to show context preview
        html_content += """
        <button class="convert-button" onclick="showContextPreview()">Show Context Preview</button>
        <script>
        function showContextPreview() {
            document.querySelectorAll('.context-preview').forEach(function(preview) {
                preview.classList.add('show-context');
            });
            document.querySelector('.convert-button').style.display = 'none';
        }
        </script>
        """
        
        # Convert table data to RAG format
        rag_content = convert_to_rag_format(table_data)
        
        # Add RAG content to the page (hidden by default)
        html_content += f"""
        <div class="context-preview" id="rag-content">
            <h3>RAG Format</h3>
            <pre style="white-space: pre-wrap; font-family: 'Poppins', monospace; font-size: 0.9em;">
{rag_content}
            </pre>
        </div>
        """
        
        # Add structured data
        structured_data = {
            "@context": "https://schema.org/",
            "@type": "ItemList",
            "itemListElement": [
                {
                    "@type": "Product",
                    "name": ctx.get('product_name', 'Unknown Product'),
                    "productNumber": ctx.get('product_number', ''),
                    "description": ctx.get('description', ''),
                    "category": ctx.get('category', ''),
                    "warranty": ctx.get('warranty', ''),
                    "offers": {
                        "@type": "Offer",
                        "price": str(ctx.get('price', '')) if ctx.get('price') is not None else '',
                        "priceCurrency": "USD"
                    }
                }
                for ctx in product_contexts
            ],
            "productGroups": product_groups
        }
        
        html_content += f"""
        <script type="application/ld+json">
            {json.dumps(structured_data)}
        </script>
        <div id="product-contexts" style="display: none;" data-context-version="1.0">
            {json.dumps({
                'products': product_contexts,
                'groups': product_groups,
                'common_questions': [
                    {
                        'question': 'What are the product numbers for {}?',
                        'type': 'product_numbers_by_name'
                    },
                    {
                        'question': 'What is the price of {}?',
                        'type': 'price_by_product_number'
                    },
                    {
                        'question': 'Can you describe {}?',
                        'type': 'description_by_product_number'
                    }
                ]
            })}
        </div>
        """
        
        logger.debug("HTML conversion completed successfully")
        return html_content
    except Exception as e:
        logger.error(f"Error converting Excel to HTML: {str(e)}")
        raise Exception(f"Error converting Excel to HTML: {str(e)}")

def publish_to_wordpress(title, content):
    try:
        logger.debug(f"Publishing to WordPress with title: {title}")
        
        # First, get the parent page ID for "Price List"
        parent_url = f"{WP_API_URL}?search=Price List"
        logger.debug(f"Searching for parent page: {parent_url}")
        
        credentials = f"{WP_USER}:{WP_APP_PASSWORD}"
        token = base64.b64encode(credentials.encode())
        headers = {
            "Authorization": f"Basic {token.decode('utf-8')}",
            "Content-Type": "application/json"
        }
        
        parent_response = requests.get(
            parent_url,
            headers=headers
        )
        
        parent_id = None
        if parent_response.status_code == 200:
            pages = parent_response.json()
            for page in pages:
                if page['title']['rendered'] == "Price List":
                    parent_id = page['id']
                    logger.debug(f"Found parent page 'Price List' with ID: {parent_id}")
                    break
        
        if not parent_id:
            logger.warning("Parent page 'Price List' not found. Creating as a top-level page.")
        
        # Prepare the page data
        page_data = {
            'title': title,
            'content': content,
            'status': 'publish',
            'parent': parent_id if parent_id else 0
        }
        
        # Make the API request
        logger.debug("Sending request to WordPress API")
        response = requests.post(
            WP_API_URL,
            headers=headers,
            json=page_data
        )
        
        logger.debug(f"WordPress API response status: {response.status_code}")
        
        if response.status_code == 201:
            page_url = response.json()['link']
            logger.info(f"Page published successfully: {page_url}")
            return page_url
        else:
            error_message = response.text
            try:
                error_json = response.json()
                if 'message' in error_json:
                    error_message = error_json['message']
            except:
                pass
            logger.error(f"WordPress API error: {error_message}")
            raise Exception(f"WordPress API error: {error_message}")
    except Exception as e:
        logger.error(f"Error publishing to WordPress: {str(e)}")
        raise Exception(f"Error publishing to WordPress: {str(e)}")

@app.route('/api/process', methods=['POST'])
def process_file():
    try:
        logger.debug("Received process request")
        if 'file' not in request.files:
            return jsonify({
                'status': 'error',
                'message': 'No file provided',
                'ui': {
                    'type': 'error',
                    'title': 'Upload Failed',
                    'description': 'Please select an Excel file to upload.'
                }
            }), 400
            
        file = request.files['file']
        action = request.form.get('action')
        
        # Get filename without extension for default title
        filename = os.path.splitext(file.filename)[0]
        # Use provided title or default to filename
        title = request.form.get('title') or filename
        
        logger.debug(f"Processing request - Action: {action}, Title: {title}, Original filename: {file.filename}")
        
        if not file.filename.endswith('.xlsx'):
            return jsonify({
                'status': 'error',
                'message': 'Invalid file type',
                'ui': {
                    'type': 'error',
                    'title': 'Invalid File',
                    'description': 'Please upload an Excel (.xlsx) file only.'
                }
            }), 400
            
        # Save the file temporarily
        temp_path = f"temp_{int(time.time())}.xlsx"
        logger.debug(f"Saving temporary file: {temp_path}")
        file.save(temp_path)
        
        try:
            if action == 'convert':
                # --- context summary extraction ---
                wb = load_workbook(temp_path, data_only=True)
                contextData = {}
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    headers = {}
                    first_row = next(ws.rows)
                    for idx, cell in enumerate(first_row):
                        header_value = get_cell_value_safe(cell)
                        if header_value:
                            headers[idx] = str(header_value).strip()
                    sheet_data = []
                    for row_idx, row in enumerate(ws.rows):
                        if row_idx == 0:
                            continue
                        row_data = {}
                        for col_idx, cell in enumerate(row):
                            if col_idx in headers:
                                row_data[headers[col_idx]] = get_cell_value_safe(cell)
                        if row_data:
                            sheet_data.append(row_data)
                    if sheet_data:
                        contextData[sheet] = generate_contexts_from_excel(sheet_data, sheet)
                # --- end context summary extraction ---
                html_content = convert_excel_to_html(temp_path)
                return jsonify({
                    'status': 'success',
                    'html': html_content,
                    'defaultTitle': filename,
                    'contextData': contextData,
                    'ui': {
                        'type': 'success',
                        'title': 'Conversion Successful',
                        'description': 'Your Excel file has been converted to HTML format.'
                    }
                })
            
            elif action == 'publish':
                if not title:
                    return jsonify({
                        'status': 'error',
                        'message': 'Title is required',
                        'ui': {
                            'type': 'error',
                            'title': 'Missing Title',
                            'description': 'Please provide a title for the price list.'
                        }
                    }), 400
                    
                html_content = convert_excel_to_html(temp_path)
                page_url = publish_to_wordpress(title, html_content)
                
                return jsonify({
                    'status': 'success',
                    'url': page_url,
                    'defaultTitle': filename,
                    'ui': {
                        'type': 'success',
                        'title': 'Published Successfully',
                        'description': f'Your price list has been published. View it <a href="{page_url}" target="_blank">here</a>.'
                    }
                })
                
            else:
                return jsonify({
                    'status': 'error',
                    'message': 'Invalid action',
                    'ui': {
                        'type': 'error',
                        'title': 'Invalid Action',
                        'description': 'The requested action is not supported.'
                    }
                }), 400
                
        finally:
            # Clean up temporary file
            if os.path.exists(temp_path):
                logger.debug(f"Cleaning up temporary file: {temp_path}")
                os.remove(temp_path)
                
    except Exception as e:
        logger.error(f"Error processing request: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e),
            'ui': {
                'type': 'error',
                'title': 'Processing Error',
                'description': f'An error occurred: {str(e)}'
            }
        }), 500

@app.route('/api/rag/convert', methods=['POST'])
def rag_convert():
    """
    Convert Excel data to RAG format with customizable options
    """
    try:
        if 'file' not in request.files:
            return jsonify({
                'status': 'error',
                'message': 'No file provided'
            }), 400
            
        file = request.files['file']
        if not file.filename.endswith('.xlsx'):
            return jsonify({
                'status': 'error',
                'message': 'Invalid file type. Please upload an Excel (.xlsx) file.'
            }), 400

        # Get optional parameters
        format_type = request.form.get('format', 'default')  # default, json, csv
        include_metadata = request.form.get('include_metadata', 'false').lower() == 'true'
        custom_prompt = request.form.get('prompt')
        use_ollama = request.form.get('use_ollama', 'true').lower() == 'true'
            
        # Save the file temporarily
        temp_path = f"temp_{int(time.time())}.xlsx"
        file.save(temp_path)
        
        try:
            # Convert to RAG format
            wb = load_workbook(temp_path, data_only=True)
            table_data = []
            metadata = {
                'total_sheets': len(wb.sheetnames),
                'sheets': [],
                'total_rows': 0,
                'total_columns': 0,
                'conversion_method': 'ollama' if use_ollama else 'simple'
            }
            
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                headers = {}
                first_row = next(ws.rows)
                
                for idx, cell in enumerate(first_row):
                    header_value = get_cell_value_safe(cell)
                    if header_value:
                        headers[idx] = str(header_value).strip()
                
                sheet_data = []
                for row_idx, row in enumerate(ws.rows):
                    if row_idx == 0:  # Skip header row
                        continue
                        
                    row_data = {}
                    for col_idx, cell in enumerate(row):
                        if col_idx in headers:
                            row_data[headers[col_idx]] = get_cell_value_safe(cell)
                    
                    if row_data:  # Only add non-empty rows
                        sheet_data.append(row_data)
                        table_data.append(row_data)
                
                metadata['sheets'].append({
                    'name': sheet,
                    'rows': len(sheet_data),
                    'columns': len(headers)
                })
                metadata['total_rows'] += len(sheet_data)
                metadata['total_columns'] = max(metadata['total_columns'], len(headers))
            
            # Convert to RAG format
            rag_content = convert_to_rag_format(table_data, use_ollama)
            
            # Format response based on requested format
            if format_type == 'json':
                # Convert RAG content to structured JSON
                items = []
                for line in rag_content.split('\n'):
                    if not line.strip():
                        continue
                    try:
                        item = line.split(' is a ')[0].strip('"')
                        description = line.split(' is a ')[1].split(' that costs ')[0].strip('"')
                        price = line.split(' that costs ')[1].strip('"')
                        items.append({
                            'item': item,
                            'description': description,
                            'price': price
                        })
                    except:
                        continue
                
                result = {
                    'items': items,
                    'metadata': metadata if include_metadata else None
                }
            elif format_type == 'csv':
                # Convert to CSV format
                csv_lines = ['Item,Description,Price']
                for line in rag_content.split('\n'):
                    if not line.strip():
                        continue
                    try:
                        item = line.split(' is a ')[0].strip('"')
                        description = line.split(' is a ')[1].split(' that costs ')[0].strip('"')
                        price = line.split(' that costs ')[1].strip('"')
                        csv_lines.append(f'"{item}","{description}","{price}"')
                    except:
                        continue
                result = '\n'.join(csv_lines)
            else:
                # Default format (text)
                result = {
                    'content': rag_content,
                    'metadata': metadata if include_metadata else None
                }
            
            return jsonify({
                'status': 'success',
                'format': format_type,
                'data': result
            })
            
        finally:
            if os.path.exists(temp_path):
                os.remove(temp_path)
                
    except Exception as e:
        logger.error(f"Error in RAG conversion: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@app.route('/api/rag/validate', methods=['POST'])
def rag_validate():
    """
    Validate if the Excel file structure is suitable for RAG conversion
    """
    try:
        if 'file' not in request.files:
            return jsonify({
                'status': 'error',
                'message': 'No file provided'
            }), 400
            
        file = request.files['file']
        if not file.filename.endswith('.xlsx'):
            return jsonify({
                'status': 'error',
                'message': 'Invalid file type'
            }), 400
            
        # Save the file temporarily
        temp_path = f"temp_{int(time.time())}.xlsx"
        file.save(temp_path)
        
        try:
            wb = load_workbook(temp_path, data_only=True)
            validation_results = {
                'is_valid': True,
                'issues': [],
                'recommendations': []
            }
            
            # Check each sheet
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                headers = {}
                first_row = next(ws.rows)
                
                # Check headers
                for idx, cell in enumerate(first_row):
                    header_value = get_cell_value_safe(cell)
                    if header_value:
                        headers[idx] = str(header_value).strip()
                
                if len(headers) < 3:
                    validation_results['is_valid'] = False
                    validation_results['issues'].append(f"Sheet '{sheet}' has less than 3 columns")
                    validation_results['recommendations'].append(f"Add more columns to sheet '{sheet}'")
                
                # Check for required column types
                has_item = False
                has_description = False
                has_price = False
                
                for header in headers.values():
                    header_lower = header.lower()
                    if any(term in header_lower for term in ['item', 'product', 'sku', 'code']):
                        has_item = True
                    if any(term in header_lower for term in ['description', 'detail', 'info']):
                        has_description = True
                    if any(term in header_lower for term in ['price', 'cost', 'amount', '$']):
                        has_price = True
                
                if not has_item:
                    validation_results['recommendations'].append(f"Add an 'Item' or 'Product' column to sheet '{sheet}'")
                if not has_description:
                    validation_results['recommendations'].append(f"Add a 'Description' column to sheet '{sheet}'")
                if not has_price:
                    validation_results['recommendations'].append(f"Add a 'Price' column to sheet '{sheet}'")
            
            return jsonify({
                'status': 'success',
                'validation': validation_results
            })
            
        finally:
            if os.path.exists(temp_path):
                os.remove(temp_path)
                
    except Exception as e:
        logger.error(f"Error in RAG validation: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

if __name__ == '__main__':
    logger.info("Starting Flask server")
    app.run(debug=True) 